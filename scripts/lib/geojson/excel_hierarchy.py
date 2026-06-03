"""Load (基地, 片区) → [地块名] from 农业资产盘点明细.xlsx."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

YOUYANG_SHEET = " 重庆酉阳"


def _rows_from_sheet(ws) -> tuple[list, dict[str, int]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}
    header = rows[0]
    col = {v: i for i, v in enumerate(header) if v}
    return rows, col


def _ingest_rows(
    rows: list,
    col: dict[str, int],
    result: dict[tuple[str, str], list[str]],
    seen: set[tuple[str, str, str]],
) -> None:
    for row in rows[1:]:
        base = row[col["基地名称"]] if "基地名称" in col else None
        district = row[col["片区名称"]] if "片区名称" in col else None
        plot = row[col["地块名称"]] if "地块名称" in col else None
        if base and district and plot:
            key = (base, district, plot)
            if key not in seen:
                result[(base, district)].append(plot)
                seen.add(key)


def load_excel_hierarchy(excel_path: Path) -> dict[tuple[str, str], list[str]]:
    try:
        import openpyxl
    except ImportError:
        print("  ⚠ openpyxl 未安装（pip install openpyxl），跳过 Excel 层级")
        return {}

    if not excel_path.is_file():
        print(f"  ⚠ Excel 不存在: {excel_path}")
        return {}

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    result: dict[tuple[str, str], list[str]] = defaultdict(list)
    seen: set[tuple[str, str, str]] = set()

    _ingest_rows(*_rows_from_sheet(wb.worksheets[0]), result, seen)
    if YOUYANG_SHEET in wb.sheetnames:
        _ingest_rows(*_rows_from_sheet(wb[YOUYANG_SHEET]), result, seen)
        print(f"  Excel: 已合并工作表「{YOUYANG_SHEET.strip()}」")
    wb.close()
    return dict(result)


def load_youyang_contract_district_map(
    excel_path: Path,
) -> dict[str, set[str]]:
    """合同编号 → 片区名称（重庆酉阳专用表）。"""
    try:
        import openpyxl
    except ImportError:
        return {}

    if not excel_path.is_file():
        return {}

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    if YOUYANG_SHEET not in wb.sheetnames:
        wb.close()
        return {}

    rows, col = _rows_from_sheet(wb[YOUYANG_SHEET])
    wb.close()
    if "合同编号" not in col or "片区名称" not in col:
        return {}

    out: dict[str, set[str]] = defaultdict(set)
    for row in rows[1:]:
        contract = row[col["合同编号"]]
        district = row[col["片区名称"]]
        if contract and district:
            out[str(contract).strip()].add(str(district).strip())
    return {k: set(v) for k, v in out.items()}


def youyang_district_names(hierarchy: dict[tuple[str, str], list[str]]) -> list[str]:
    return sorted(
        district
        for (base, district) in hierarchy
        if base == "重庆酉阳"
    )
